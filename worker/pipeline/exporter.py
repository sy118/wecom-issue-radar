from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Iterable


CHAT_HEADERS = [
    "日期",
    "时间",
    "群名称",
    "发送人",
    "消息类型",
    "消息内容",
    "OCR文字",
    "图片数量",
    "附件路径",
    "消息ID",
]


def export_day(
    day_dir: str | Path,
    date_text: str,
    group_name: str,
    *,
    export_xlsx: bool = True,
    export_markdown: bool = True,
    include_issues: bool = True,
    output_dir: str | Path | None = None,
) -> dict[str, str]:
    day_path = Path(day_dir)
    destination = Path(output_dir) if output_dir else day_path / "exports"
    destination.mkdir(parents=True, exist_ok=True)
    messages = read_jsonl(day_path / "raw_messages.jsonl")
    ocr_map = load_ocr(day_path / "grouped_issues" / "image_ocr.jsonl")
    issues = load_issues(day_path, date_text) if include_issues else []
    rows = build_chat_rows(messages, ocr_map, group_name)
    safe_group = safe_filename(group_name or "企业微信群")
    stem = f"{date_text}_{safe_group}_聊天与问题盘点"
    outputs: dict[str, str] = {}
    if export_xlsx:
        xlsx_path = destination / f"{stem}.xlsx"
        write_xlsx(xlsx_path, rows, issues, date_text, group_name)
        outputs["xlsx"] = str(xlsx_path)
    if export_markdown:
        md_path = destination / f"{stem}.md"
        write_markdown(md_path, messages, ocr_map, issues, date_text, group_name)
        outputs["markdown"] = str(md_path)
    return outputs


def read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"聊天记录不存在: {path}")
    records = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            if line.strip():
                records.append(json.loads(line))
    return sorted(records, key=lambda row: (int(row.get("send_time") or 0), int(row.get("message_id") or 0)))


def load_ocr(path: Path) -> dict[int, str]:
    if not path.exists():
        return {}
    result: dict[int, list[str]] = {}
    for row in read_jsonl(path):
        message_id = int(row.get("source_message_id") or 0)
        text = str(row.get("ocr_text") or "").strip()
        if message_id and text:
            result.setdefault(message_id, []).append(text)
    return {key: "\n".join(values) for key, values in result.items()}


def load_issues(day_dir: Path, date_text: str) -> list[dict]:
    ymd = date_text.replace("-", "")
    candidates = [
        day_dir / "grouped_issues" / f"issue_definitions_{ymd}.json",
        day_dir / "grouped_issues" / f"final_issues_{ymd}.json",
    ]
    for path in candidates:
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8") as file:
            data = json.load(file)
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and isinstance(data.get("issues"), list):
            return data["issues"]
    return []


def build_chat_rows(messages: Iterable[dict], ocr_map: dict[int, str], group_name: str) -> list[list[object]]:
    rows = []
    for message in messages:
        message_id = int(message.get("message_id") or 0)
        files = message.get("files") or []
        attachment_paths = [str(item.get("local_path") or "") for item in files if item.get("local_path")]
        rows.append(
            [
                str(message.get("date") or ""),
                str(message.get("message_time") or ""),
                str(message.get("conversation_name") or group_name or ""),
                str(message.get("sender") or ""),
                str(message.get("type") or message.get("content_type") or ""),
                str(message.get("raw_text") or ""),
                ocr_map.get(message_id, ""),
                int(message.get("image_count_visible") or 0),
                "\n".join(attachment_paths),
                message_id,
            ]
        )
    return rows


def write_xlsx(path: Path, rows: list[list[object]], issues: list[dict], date_text: str, group_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("导出 Excel 需要 openpyxl，请先执行 uv sync") from exc

    workbook = Workbook()
    info = workbook.active
    info.title = "导出说明"
    info_rows = [
        ("导出日期", date_text),
        ("群名称", group_name),
        ("生成时间", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %z")),
        ("聊天条数", len(rows)),
        ("问题条数", len(issues)),
        ("说明", "聊天记录为本地企微数据导出；OCR 文字仅在启用 OCR 时存在。"),
    ]
    for key, value in info_rows:
        info.append([key, value])
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 80

    chat = workbook.create_sheet("聊天记录")
    chat.append(CHAT_HEADERS)
    for row in rows:
        chat.append(row)
    style_sheet(chat, [12, 20, 24, 16, 14, 60, 60, 12, 50, 16], get_column_letter, Font, PatternFill, Alignment)

    issue_sheet = workbook.create_sheet("问题清单")
    issue_headers = ["序号", "时间", "发送人", "模块", "问题分类", "问题描述", "问题总结", "原因/结论", "截图引用", "问题Key"]
    issue_sheet.append(issue_headers)
    for index, issue in enumerate(issues, start=1):
        issue_sheet.append(
            [
                index,
                issue.get("message_time") or "",
                issue.get("sender") or "",
                issue.get("module_text") or issue.get("模块") or "",
                issue.get("issue_category_text") or issue.get("问题分类") or "",
                issue.get("problem_description") or issue.get("问题描述") or "",
                issue.get("issue_summary_text") or issue.get("问题总结") or "",
                issue.get("reason") or issue.get("原因") or "",
                "\n".join(issue.get("image_refs") or issue.get("问题截图") or []),
                issue.get("key") or "",
            ]
        )
    style_sheet(issue_sheet, [8, 20, 16, 16, 18, 50, 30, 70, 40, 28], get_column_letter, Font, PatternFill, Alignment)
    workbook.save(path)


def style_sheet(sheet, widths, get_column_letter, Font, PatternFill, Alignment) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_markdown(
    path: Path,
    messages: list[dict],
    ocr_map: dict[int, str],
    issues: list[dict],
    date_text: str,
    group_name: str,
) -> None:
    lines = [
        f"# {date_text} {group_name or '企业微信群'}聊天与问题盘点",
        "",
        f"- 导出时间：{datetime.now().astimezone().strftime('%Y-%m-%d %H:%M:%S %z')}",
        f"- 聊天条数：{len(messages)}",
        f"- 问题条数：{len(issues)}",
        "",
    ]
    if issues:
        lines.extend(["## 问题清单", ""])
        for index, issue in enumerate(issues, start=1):
            summary = issue.get("issue_summary_text") or issue.get("问题总结") or issue.get("problem_description") or issue.get("问题描述") or "未命名问题"
            lines.extend(
                [
                    f"### {index}. {summary}",
                    "",
                    f"- 模块：{issue.get('module_text') or issue.get('模块') or '待评估'}",
                    f"- 分类：{issue.get('issue_category_text') or issue.get('问题分类') or '待评估'}",
                    f"- 时间：{issue.get('message_time') or ''}",
                    f"- 描述：{issue.get('problem_description') or issue.get('问题描述') or ''}",
                    "",
                    str(issue.get("reason") or issue.get("原因") or ""),
                    "",
                ]
            )

    lines.extend(["## 完整聊天记录", ""])
    for message in messages:
        message_id = int(message.get("message_id") or 0)
        time_text = str(message.get("message_time") or "")
        sender = str(message.get("sender") or "未知")
        raw_text = str(message.get("raw_text") or "").strip() or "（无文字消息）"
        lines.extend([f"### {time_text} · {sender}", "", raw_text, ""])
        ocr_text = ocr_map.get(message_id)
        if ocr_text:
            lines.extend(["> 截图 OCR：", ">", *[f"> {line}" for line in ocr_text.splitlines()], ""])
        image_paths = message.get("image_paths") or []
        if image_paths:
            lines.extend(["附件：", "", *[f"- `{item}`" for item in image_paths], ""])
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def safe_filename(value: str) -> str:
    text = "".join("_" if char in '<>:"/\\|?*\r\n\t' else char for char in str(value or ""))
    return text.strip(" .")[:80] or "企业微信群"

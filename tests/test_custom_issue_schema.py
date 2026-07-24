from __future__ import annotations

import copy
import hashlib
import json
import multiprocessing
import os
import struct
import tempfile
import textwrap
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from worker.pipeline.config_store import (
    load_config,
    save_config,
    selected_prompt,
    selected_smart_sheet_template,
)
from worker.pipeline.exporter import export_day
from worker.pipeline.issue_schema import (
    FIELD_TYPES,
    build_model_contract,
    display_value,
    issue_value,
    normalize_field_value,
    normalize_issue_fields,
    normalize_issue_values,
    validate_issue_fields,
)
from worker.pipeline.llm_analyzer import analyze_day, build_instruction, build_issue_definitions
from worker.pipeline.smart_sheet import (
    date_time_millis,
    document_revision,
    issue_dedupe_identity,
    issue_values,
    load_ledger,
    preview_sync,
    resolve_template,
    resolve_webhook_url,
    save_ledger,
    sync_issues,
    template_revision,
    template_synced,
    utc_date_millis,
)


DATE_TEXT = "2026-07-24"


def canonical_field(
    key: str,
    label: str,
    field_type: str,
    *,
    required: bool = False,
    options: list[str] | None = None,
    default_value="",
    instruction: str = "",
) -> dict:
    return {
        "key": key,
        "label": label,
        "type": field_type,
        "required": required,
        "instruction": instruction,
        "options": options or [],
        "default_value": default_value,
    }


def write_issue_document(
    day_dir: Path,
    *,
    issue_fields: list[dict],
    issues: list[dict],
    prompt_template_id: str = "",
) -> Path:
    grouped = day_dir / "grouped_issues"
    grouped.mkdir(parents=True, exist_ok=True)
    snapshot_dir = grouped / "snapshots"
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    path = snapshot_dir / "issue_definitions_20260724_test_snapshot.json"
    serialized = json.dumps(
        {
            "schema_version": 2,
            "date": DATE_TEXT,
            "prompt": {
                "id": "custom_prompt",
                "name": "自定义提示词",
                "default_smart_sheet_template_id": prompt_template_id,
            },
            "issue_fields": issue_fields,
            "image_manifest": {},
            "issues": issues,
        },
        ensure_ascii=False,
    )
    path.write_text(serialized, encoding="utf-8")
    (grouped / "issue_definitions_20260724.json").write_text(
        serialized,
        encoding="utf-8",
    )
    return path


def test_definition_path(day_dir: Path) -> Path:
    return (
        day_dir
        / "grouped_issues"
        / "snapshots"
        / "issue_definitions_20260724_test_snapshot.json"
    )


def set_test_image_manifest(day_dir: Path, image_manifest: dict[str, str]) -> None:
    paths = [
        test_definition_path(day_dir),
        day_dir / "grouped_issues" / "issue_definitions_20260724.json",
    ]
    for path in paths:
        document = json.loads(path.read_text(encoding="utf-8"))
        document["image_manifest"] = image_manifest
        path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")


def smart_template(
    template_id: str,
    *,
    webhook_url: str = "https://example.invalid/hook",
    webhook_url_env: str = "",
) -> dict:
    return {
        "id": template_id,
        "name": f"模板 {template_id}",
        "url": f"https://docs.qq.com/sheet/{template_id}",
        "webhook_url": webhook_url,
        "webhook_url_env": webhook_url_env,
        "batch_size": 20,
        "schema": {
            "f_desc": {"title": "*问题描述", "type": "text"},
        },
        "field_mappings": [
            {
                "source_key": "description",
                "target_field_id": "f_desc",
                "target_type": "text",
                "required": True,
                "default_value": "",
            }
        ],
    }


def run_sync_process(
    role: str,
    template_id: str,
    fail_first: bool,
    config: dict,
    day_dir: str,
    ready,
    entered,
    release,
    second_posted,
    results,
) -> None:
    def fake_post(_url, _payload):
        if role == "first":
            entered.set()
            if not release.wait(10):
                raise RuntimeError("timed out waiting to release the first sync")
            if fail_first:
                raise RuntimeError("intentional first-sync failure")
        else:
            second_posted.set()
        return {
            "errcode": 0,
            "add_records": [{"record_id": f"remote-record-{template_id}"}],
        }

    try:
        ready.set()
        with patch("worker.pipeline.smart_sheet.post_json", side_effect=fake_post):
            result = sync_issues(
                config,
                day_dir,
                DATE_TEXT,
                template_id=template_id,
                upload_images=False,
                definition_path=test_definition_path(Path(day_dir)),
            )
    except Exception as exc:  # pragma: no cover - asserted in the parent process
        results.put({"role": role, "error": f"{type(exc).__name__}: {exc}"})
    else:
        results.put({"role": role, "result": result})


class IssueSchemaTests(unittest.TestCase):
    def test_model_contract_uses_only_explicit_defaults_as_examples(self):
        fields = [
            canonical_field(
                "severity",
                "严重级别",
                "single_select",
                required=True,
                options=["P0", "P1"],
            ),
            canonical_field("resolved", "已恢复", "boolean", required=True),
            canonical_field("affected", "影响数", "number", required=True),
            canonical_field("tags", "标签", "multiple_select", options=["线上"]),
            canonical_field(
                "priority",
                "优先级",
                "single_select",
                options=["高", "低"],
                default_value="低",
            ),
        ]
        contract = build_model_contract(fields)
        example = json.loads(contract.split("结构必须是：\n", 1)[1].split("\n没有问题时", 1)[0])

        self.assertEqual(
            example["issues"][0]["values"],
            {
                "severity": None,
                "resolved": None,
                "affected": None,
                "tags": [],
                "priority": "低",
            },
        )
        self.assertNotIn('"severity": "P0"', contract)

    def test_strict_validation_accepts_supported_types_and_rejects_bad_schema(self):
        fields = [
            canonical_field(f"field_{index}", field_type, field_type)
            for index, field_type in enumerate(sorted(FIELD_TYPES), start=1)
        ]
        fields.append(canonical_field("legacy_multi", "兼容多选", "multi_select"))
        validate_issue_fields(fields)
        self.assertEqual(
            normalize_issue_fields([fields[-1]])[0]["type"],
            "multiple_select",
        )

        invalid_schemas = {
            "empty": [],
            "non-dict item": ["bad"],
            "invalid key": [canonical_field("bad-key", "坏 Key", "text")],
            "reserved key": [canonical_field("values", "系统字段", "text")],
            "duplicate key": [
                canonical_field("same", "一", "text"),
                canonical_field("same", "二", "text"),
            ],
            "unknown type": [canonical_field("field", "字段", "currency")],
        }
        for label, schema in invalid_schemas.items():
            with self.subTest(label=label), self.assertRaises(ValueError):
                validate_issue_fields(schema)

    def test_field_types_are_normalized_without_losing_zero_or_false(self):
        cases = [
            (
                canonical_field(
                    "severity",
                    "严重级别",
                    "single_select",
                    options=["P0", "P1"],
                    default_value="P1",
                ),
                "not-an-option",
                "P1",
            ),
            (
                canonical_field(
                    "tags",
                    "标签",
                    "multiple_select",
                    options=["客户", "线上"],
                ),
                ["客户", "无效", "客户", "线上"],
                ["客户", "线上"],
            ),
            (canonical_field("resolved", "已恢复", "boolean"), False, False),
            (canonical_field("affected", "影响数", "number"), 0, 0),
            (canonical_field("ratio", "比例", "number"), "2.5", 2.5),
            (canonical_field("day", "日期", "date"), "2026-07-24 18:00", "2026-07-24"),
            (
                canonical_field("at", "时间", "datetime"),
                "2026-07-24T18:00:00+08:00",
                "2026-07-24T18:00:00+08:00",
            ),
            (
                canonical_field("link", "链接", "url"),
                "https://example.com/ticket/1",
                "https://example.com/ticket/1",
            ),
            (canonical_field("bad_link", "坏链接", "url"), "ftp://example.com", ""),
        ]
        for field, value, expected in cases:
            with self.subTest(field=field["key"]):
                self.assertEqual(normalize_field_value(value, field), expected)

    def test_required_single_select_without_default_does_not_invent_an_option(self):
        field = canonical_field(
            "severity",
            "严重级别",
            "single_select",
            required=True,
            options=["P0", "P1"],
        )

        self.assertEqual(normalize_field_value(None, field), "")
        self.assertEqual(normalize_field_value("not-an-option", field), "")

    def test_required_boolean_without_default_does_not_invent_false(self):
        field = canonical_field("resolved", "已恢复", "boolean", required=True)
        false_default = canonical_field(
            "resolved",
            "已恢复",
            "boolean",
            required=True,
            default_value=False,
        )

        self.assertEqual(normalize_field_value(None, field), "")
        self.assertEqual(normalize_field_value("unknown", field), "")
        self.assertIs(normalize_field_value(False, field), False)
        self.assertIs(normalize_field_value(None, false_default), False)

    def test_nested_values_win_over_legacy_flat_values_including_zero_and_false(self):
        fields = [
            canonical_field("description", "描述", "long_text"),
            canonical_field("affected", "影响数", "number", default_value=99),
            canonical_field("resolved", "已恢复", "boolean", default_value=True),
            canonical_field("legacy_only", "旧字段", "text"),
        ]
        raw_issue = {
            "values": {
                "description": "嵌套值",
                "affected": 0,
                "resolved": False,
            },
            "description": "旧扁平值",
            "affected": 88,
            "resolved": True,
            "legacy_only": "仍兼容扁平格式",
        }

        self.assertEqual(
            normalize_issue_values(raw_issue, fields),
            {
                "description": "嵌套值",
                "affected": 0,
                "resolved": False,
                "legacy_only": "仍兼容扁平格式",
            },
        )
        self.assertEqual(issue_value(raw_issue, "affected", 123), 0)
        self.assertIs(issue_value(raw_issue, "resolved", True), False)
        self.assertEqual(issue_value(raw_issue, "legacy_only"), "仍兼容扁平格式")
        self.assertEqual(display_value(0), "0")
        self.assertEqual(display_value(False), "否")

    def test_legacy_chinese_flat_keys_are_resolved_only_as_fallbacks(self):
        legacy_issue = {
            "模块": "订单/售后",
            "问题分类": "功能缺陷",
            "问题描述": "支付按钮无响应",
            "问题总结": "支付失败",
            "原因": "前端请求未发出",
        }
        expected = {
            "module_text": "订单/售后",
            "issue_category_text": "功能缺陷",
            "problem_description": "支付按钮无响应",
            "issue_summary_text": "支付失败",
            "reason": "前端请求未发出",
        }

        for key, value in expected.items():
            with self.subTest(key=key):
                self.assertEqual(issue_value(legacy_issue, key), value)

        modern_issue = {
            **legacy_issue,
            "problem_description": "英文扁平键",
            "values": {"problem_description": "新 schema 值"},
        }
        self.assertEqual(
            issue_value(modern_issue, "problem_description"),
            "新 schema 值",
        )


class ConfigMigrationTests(unittest.TestCase):
    def test_legacy_config_migrates_and_save_load_is_idempotent(self):
        custom_fields = [
            canonical_field(
                "severity",
                "严重级别",
                "single_select",
                required=True,
                options=["S0", "S1", "S2"],
                default_value="S1",
                instruction="按影响范围判断",
            ),
            canonical_field("affected", "影响数", "number", default_value=0),
            canonical_field("resolved", "是否恢复", "boolean", default_value=False),
        ]
        legacy = {
            "config_version": 1,
            "prompts": {
                "default_id": "incident_custom",
                "items": [
                    {
                        "id": "incident_custom",
                        "name": "自定义故障",
                        "description": "测试迁移",
                        "content": "识别故障",
                        "issue_fields": custom_fields,
                        "default_smart_sheet_template_id": "default",
                    }
                ],
            },
            "smart_sheet": {
                "url": "https://docs.qq.com/sheet/legacy",
                "webhook_url_env": "LEGACY_SMART_SHEET_HOOK",
                "webhook_url": "https://example.invalid/legacy-hook",
                "batch_size": 13,
                "schema": {
                    "f04Gwj": {
                        "title": "*模块",
                        "type": "single_select",
                        "enum": ["自定义模块 A", "自定义模块 B"],
                    },
                    "ftk5Tx": {"title": "*问题描述", "type": "text"},
                    "fsFBqK": {
                        "title": "问题分类",
                        "type": "single_select",
                        "enum": ["自定义分类", "待评估"],
                    },
                },
                "defaults": {
                    "module_text": "自定义模块 B",
                    "issue_category_text": "自定义分类",
                },
            },
        }

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "config.local.json"
            path.write_text(json.dumps(legacy, ensure_ascii=False), encoding="utf-8")

            first, _ = load_config(path)
            self.assertEqual(first["config_version"], 2)
            template = first["smart_sheet"]["templates"][0]
            self.assertEqual(template["id"], "default")
            self.assertEqual(template["url"], legacy["smart_sheet"]["url"])
            self.assertEqual(template["webhook_url"], legacy["smart_sheet"]["webhook_url"])
            self.assertEqual(template["batch_size"], 13)
            self.assertEqual(
                template["schema"]["f04Gwj"]["enum"],
                ["自定义模块 A", "自定义模块 B"],
            )
            prompt = first["prompts"]["items"][0]
            self.assertEqual(prompt["issue_fields"], custom_fields)
            self.assertEqual(prompt["default_smart_sheet_template_id"], "default")

            save_config(first, path)
            persisted = json.loads(path.read_text(encoding="utf-8"))
            for legacy_key in (
                "url",
                "webhook_url_env",
                "webhook_url",
                "batch_size",
                "schema",
                "defaults",
            ):
                self.assertNotIn(legacy_key, persisted["smart_sheet"])

            second, _ = load_config(path)
            save_config(second, path)
            third, _ = load_config(path)
            self.assertEqual(second["prompts"], third["prompts"])
            self.assertEqual(second["smart_sheet"], third["smart_sheet"])
            self.assertEqual(
                second["prompts"]["items"][0]["issue_fields"][0]["options"],
                ["S0", "S1", "S2"],
            )
            self.assertEqual(
                second["prompts"]["items"][0]["default_smart_sheet_template_id"],
                "default",
            )

    def test_v2_preserves_explicit_empty_template_and_custom_default_fields(self):
        custom_fields = [
            canonical_field(
                "customer_impact",
                "客户影响",
                "single_select",
                options=["高", "中", "低"],
                default_value="中",
            )
        ]
        current = {
            "config_version": 2,
            "prompts": {
                "default_id": "Case Review",
                "default_issue_fields": custom_fields,
                "items": [
                    {
                        "id": "Case Review",
                        "name": "案例复盘",
                        "description": "",
                        "content": "提取客户影响",
                        "issue_fields": custom_fields,
                        "default_smart_sheet_template_id": "Incident Sheet",
                    }
                ],
            },
            "smart_sheet": {
                "default_template_id": "Incident Sheet",
                "templates": [
                    {
                        "id": "Incident Sheet",
                        "name": "待配置模板",
                        "url": "",
                        "webhook_url_env": "",
                        "webhook_url": "",
                        "batch_size": 50,
                        "schema": {},
                        "field_mappings": [],
                    }
                ],
            },
        }

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "config.local.json"
            path.write_text(json.dumps(current, ensure_ascii=False), encoding="utf-8")

            first, _ = load_config(path)
            self.assertEqual(first["prompts"]["default_issue_fields"], custom_fields)
            template = first["smart_sheet"]["templates"][0]
            self.assertEqual(template["id"], "Incident Sheet")
            self.assertEqual(template["schema"], {})
            self.assertEqual(template["field_mappings"], [])
            self.assertEqual(selected_prompt(first, "Case Review")["id"], "Case Review")
            self.assertEqual(
                selected_smart_sheet_template(first, "Incident Sheet")["id"],
                "Incident Sheet",
            )

            save_config(first, path)
            second, _ = load_config(path)
            self.assertEqual(second["prompts"]["default_issue_fields"], custom_fields)
            self.assertEqual(second["smart_sheet"]["templates"][0]["schema"], {})
            self.assertEqual(second["smart_sheet"]["templates"][0]["field_mappings"], [])

    def test_duplicate_explicit_ids_are_rejected_instead_of_rewritten(self):
        duplicate_prompts = {
            "config_version": 2,
            "prompts": {
                "items": [
                    {"id": "Same ID", "name": "A", "content": "A"},
                    {"id": "Same ID", "name": "B", "content": "B"},
                ]
            },
        }
        with self.assertRaisesRegex(ValueError, "提示词 ID 不能重复"):
            selected_prompt(duplicate_prompts)

        duplicate_templates = {
            "config_version": 2,
            "smart_sheet": {
                "templates": [
                    {"id": "Same Sheet", "name": "A"},
                    {"id": "Same Sheet", "name": "B"},
                ]
            },
        }
        with self.assertRaisesRegex(ValueError, "腾讯文档模板 ID 不能重复"):
            selected_smart_sheet_template(duplicate_templates)

    def test_ids_trim_outer_whitespace_and_missing_ids_use_stable_fallbacks(self):
        config = {
            "config_version": 2,
            "prompts": {
                "default_id": "  Case Review  ",
                "items": [
                    {
                        "id": "  Case Review  ",
                        "name": "案例复盘",
                        "content": "分析问题",
                        "default_smart_sheet_template_id": "  Incident Sheet  ",
                    },
                    {"name": "无 ID 提示词", "content": "分析问题"},
                ],
            },
            "smart_sheet": {
                "default_template_id": "  Incident Sheet  ",
                "templates": [
                    {"id": "  Incident Sheet  ", "name": "故障模板"},
                    {"name": "无 ID 模板"},
                ],
            },
        }

        prompt = selected_prompt(config, "  Case Review  ")
        template = selected_smart_sheet_template(config, "  Incident Sheet  ")

        self.assertEqual(prompt["id"], "Case Review")
        self.assertEqual(prompt["default_smart_sheet_template_id"], "Incident Sheet")
        self.assertEqual(config["prompts"]["items"][1]["id"], "prompt_2")
        self.assertEqual(config["prompts"]["default_id"], "Case Review")
        self.assertEqual(template["id"], "Incident Sheet")
        self.assertEqual(config["smart_sheet"]["templates"][1]["id"], "template_2")
        self.assertEqual(config["smart_sheet"]["default_template_id"], "Incident Sheet")
        with self.assertRaisesRegex(ValueError, "找不到提示词"):
            selected_prompt(config, "Missing Prompt")


class DynamicAnalyzerTests(unittest.TestCase):
    def test_build_instruction_snapshot_uses_only_the_selected_prompt_schema(self):
        fields = [
            canonical_field(
                "severity",
                "严重级别",
                "single_select",
                required=True,
                options=["P0", "P1"],
                default_value="P1",
                instruction="只能选择已定义级别",
            ),
            canonical_field("resolved", "是否恢复", "boolean"),
        ]
        actual = build_instruction(
            "Review {date} for {group_name}",
            DATE_TEXT,
            "Incident Group",
            {},
            issue_fields=fields,
            batch_index=2,
            batch_count=3,
        )
        expected = textwrap.dedent(
            """\
            Review 2026-07-24 for Incident Group

            分析范围：2026-07-24，群：Incident Group。
            当前为第 2/3 批；只根据本批提供的消息判断，不得编造。

            业务字段定义：
            [
              {
                "key": "severity",
                "label": "严重级别",
                "type": "single_select",
                "required": true,
                "options": [
                  "P0",
                  "P1"
                ],
                "instruction": "只能选择已定义级别"
              },
              {
                "key": "resolved",
                "label": "是否恢复",
                "type": "boolean",
                "required": false
              }
            ]

            只返回 JSON，不要 Markdown 代码块，结构必须是：
            {
              "issues": [
                {
                  "seed_message_id": 123,
                  "context_message_ids": [123, 124],
                  "question_message_ids": [123],
                  "values": {"severity": "P1", "resolved": null}
                }
              ]
            }
            没有问题时返回 {"issues": []}。seed_message_id、context_message_ids 和 question_message_ids 必须来自输入消息；values 只能使用上述字段 Key。"""
        )
        self.assertEqual(actual, expected)
        self.assertNotIn("module_text", actual)
        self.assertNotIn("issue_category_text", actual)

    def test_build_issue_definitions_snapshot_preserves_dynamic_values_and_binding(self):
        fields = [
            canonical_field("problem_description", "问题描述", "long_text", required=True),
            canonical_field(
                "severity",
                "严重级别",
                "single_select",
                required=True,
                options=["P0", "P1"],
                default_value="P1",
            ),
            canonical_field("affected", "影响数", "number", default_value=99),
            canonical_field("resolved", "已恢复", "boolean", default_value=True),
        ]
        message = {
            "message_id": 101,
            "message_time": "2026-07-24 09:30:00",
            "send_time": 123,
            "sender": "张三",
            "sender_id": 42,
            "raw_text": "登录失败",
            "dedupe_key": "R:7:101",
            "image_count_visible": 0,
        }
        prompt = {
            "id": "incident_custom",
            "name": "自定义故障",
            "issue_fields": fields,
            "default_smart_sheet_template_id": "incident_sheet",
        }
        model_issue = {
            "seed_message_id": 101,
            "context_message_ids": [101],
            "values": {
                "problem_description": "嵌套的问题描述",
                "severity": "P0",
                "affected": 0,
                "resolved": False,
            },
            "problem_description": "旧扁平描述不应覆盖嵌套值",
            "affected": 88,
            "resolved": True,
        }

        with tempfile.TemporaryDirectory() as directory:
            actual = build_issue_definitions(
                messages=[message],
                model_issues=[model_issue],
                day_dir=Path(directory),
                date_text=DATE_TEXT,
                prompt=prompt,
                config={},
            )

        actual.pop("generated_at")
        issue_key = "issue_001_" + hashlib.sha256(b"R:7:101").hexdigest()[:12]
        self.assertEqual(
            actual,
            {
                "date": DATE_TEXT,
                "range": {
                    "startDate": DATE_TEXT,
                    "startTime": "00:00",
                    "endDate": DATE_TEXT,
                    "endTime": "23:59",
                },
                "schema_version": 2,
                "generated_by": "configured_llm",
                "prompt": {
                    "id": "incident_custom",
                    "name": "自定义故障",
                    "default_smart_sheet_template_id": "incident_sheet",
                },
                "issue_fields": fields,
                "image_manifest": {},
                "issues": [
                    {
                        "key": issue_key,
                        "tenant": "",
                        "sender": "张三",
                        "sender_id": 42,
                        "message_time": "2026-07-24 09:30:00",
                        "values": {
                            "problem_description": "嵌套的问题描述",
                            "severity": "P0",
                            "affected": 0,
                            "resolved": False,
                        },
                        "module_inference": "",
                        "raw_message_keys": ["R:7:101"],
                        "context_message_keys": ["R:7:101"],
                        "expected_image_count": 0,
                        "image_refs": [],
                        "image_assignments": [],
                        "image_status": "not_required",
                        "missing_image_names": [],
                        "timeline": [
                            {
                                "message_time": "2026-07-24 09:30:00",
                                "send_time": 123,
                                "sender": "张三",
                                "sender_id": 42,
                                "role": "question",
                                "raw_text": "登录失败",
                                "raw_message_key": "R:7:101",
                                "message_id": 101,
                                "server_id": 0,
                                "image_refs": [],
                                "image_count_visible": 0,
                            }
                        ],
                        "problem_description": "嵌套的问题描述",
                    }
                ],
            },
        )

    def test_two_same_day_analyses_keep_the_first_pending_snapshot_immutable(self):
        fields = [
            canonical_field(
                "problem_description",
                "问题描述",
                "long_text",
                required=True,
            )
        ]
        template = smart_template("incident")
        template["field_mappings"][0]["source_key"] = "problem_description"
        config = {
            "prompts": {
                "default_id": "snapshot_prompt",
                "items": [
                    {
                        "id": "snapshot_prompt",
                        "name": "快照提示词",
                        "content": "识别问题",
                        "issue_fields": fields,
                        "default_smart_sheet_template_id": "incident",
                    }
                ],
            },
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            },
            "llm": {},
        }
        message = {
            "date": DATE_TEXT,
            "message_id": 101,
            "message_time": "2026-07-24 09:30:00",
            "send_time": 123,
            "sender": "张三",
            "raw_text": "登录失败",
            "dedupe_key": "R:7:101",
            "image_count_visible": 0,
        }

        def response(description: str) -> str:
            return json.dumps(
                {
                    "issues": [
                        {
                            "seed_message_id": 101,
                            "context_message_ids": [101],
                            "question_message_ids": [101],
                            "values": {"problem_description": description},
                        }
                    ]
                },
                ensure_ascii=False,
            )

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            (day_dir / "raw_messages.jsonl").write_text(
                json.dumps(message, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
            with patch(
                "worker.pipeline.llm_analyzer.call_model",
                side_effect=[response("第一份问题"), response("第二份问题")],
            ):
                first_snapshot = analyze_day(config, day_dir, DATE_TEXT, "测试群")
                second_snapshot = analyze_day(config, day_dir, DATE_TEXT, "测试群")

            self.assertNotEqual(first_snapshot, second_snapshot)
            self.assertEqual(first_snapshot.parent.name, "snapshots")
            self.assertTrue(first_snapshot.exists())
            self.assertTrue(second_snapshot.exists())
            first_document = json.loads(first_snapshot.read_text(encoding="utf-8"))
            second_document = json.loads(second_snapshot.read_text(encoding="utf-8"))
            canonical = json.loads(
                (day_dir / "grouped_issues" / "issue_definitions_20260724.json").read_text(
                    encoding="utf-8"
                )
            )
            self.assertEqual(
                first_document["issues"][0]["values"]["problem_description"],
                "第一份问题",
            )
            self.assertEqual(
                second_document["issues"][0]["values"]["problem_description"],
                "第二份问题",
            )
            self.assertEqual(canonical, second_document)

            first_preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=first_snapshot,
            )
            second_preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=second_snapshot,
            )
            self.assertEqual(first_preview["pending"], 1)
            self.assertNotEqual(
                first_preview["document_revision"],
                second_preview["document_revision"],
            )
            self.assertEqual(
                Path(first_preview["definition_path"]),
                first_snapshot.resolve(),
            )
            with patch(
                "worker.pipeline.smart_sheet.post_json",
                return_value={
                    "errcode": 0,
                    "add_records": [{"record_id": "remote-first"}],
                },
            ) as post:
                result = sync_issues(
                    config,
                    day_dir,
                    DATE_TEXT,
                    upload_images=False,
                    definition_path=first_snapshot,
                    expected_document_revision=first_preview["document_revision"],
                )

            self.assertEqual(result["synced"], 1)
            self.assertEqual(
                post.call_args.args[1]["add_records"][0]["values"]["f_desc"],
                "第一份问题",
            )

    def test_build_issue_definitions_validates_required_model_values(self):
        message = {
            "message_id": 101,
            "message_time": "2026-07-24 09:30:00",
            "send_time": 123,
            "sender": "张三",
            "raw_text": "登录失败",
            "dedupe_key": "R:7:101",
            "image_count_visible": 0,
        }

        def build(fields: list[dict], values: dict) -> dict:
            with tempfile.TemporaryDirectory() as directory:
                return build_issue_definitions(
                    messages=[message],
                    model_issues=[
                        {
                            "seed_message_id": 101,
                            "context_message_ids": [101],
                            "values": values,
                        }
                    ],
                    day_dir=Path(directory),
                    date_text=DATE_TEXT,
                    prompt={"id": "required", "name": "必填校验", "issue_fields": fields},
                    config={},
                )

        required_text = canonical_field(
            "diagnosis",
            "诊断结论",
            "text",
            required=True,
        )
        required_select = canonical_field(
            "severity",
            "严重级别",
            "single_select",
            required=True,
            options=["P0", "P1"],
        )
        with self.assertRaisesRegex(ValueError, r"问题.*101.*必填字段.*诊断结论"):
            build([required_text, required_select], {"severity": "P0"})
        for invalid in (None, "P9"):
            with self.subTest(invalid=invalid), self.assertRaisesRegex(
                ValueError,
                r"问题.*101.*必填字段.*严重级别",
            ):
                build(
                    [required_text, required_select],
                    {"diagnosis": "前端请求失败", "severity": invalid},
                )

        required_boolean = canonical_field(
            "resolved",
            "是否恢复",
            "boolean",
            required=True,
        )
        for invalid in (None, "unknown"):
            with self.subTest(boolean=invalid), self.assertRaisesRegex(
                ValueError,
                r"问题.*101.*必填字段.*是否恢复",
            ):
                build(
                    [required_text, required_boolean],
                    {"diagnosis": "前端请求失败", "resolved": invalid},
                )

        select_with_default = canonical_field(
            "severity",
            "严重级别",
            "single_select",
            required=True,
            options=["P0", "P1"],
            default_value="P1",
        )
        built = build(
            [required_text, select_with_default],
            {"diagnosis": "前端请求失败"},
        )
        self.assertEqual(built["issues"][0]["values"]["severity"], "P1")
        boolean_with_default = canonical_field(
            "resolved",
            "是否恢复",
            "boolean",
            required=True,
            default_value=False,
        )
        built = build(
            [required_text, boolean_with_default],
            {"diagnosis": "前端请求失败"},
        )
        self.assertIs(built["issues"][0]["values"]["resolved"], False)


class DynamicExportTests(unittest.TestCase):
    def test_excel_and_markdown_use_dynamic_headers_order_and_typed_values(self):
        fields = [
            canonical_field("severity", "严重级别", "single_select", options=["P0", "P1"]),
            canonical_field("tags", "标签", "multiple_select", options=["客户", "线上"]),
            canonical_field("affected", "影响数量", "number"),
            canonical_field("resolved", "已恢复", "boolean"),
        ]
        issue = {
            "key": "issue-custom-1",
            "message_time": "2026-07-24 10:00:00",
            "sender": "李四",
            "values": {
                "severity": "P0",
                "tags": ["客户", "线上"],
                "affected": 0,
                "resolved": False,
            },
            "image_refs": ["bulk:101_01"],
        }
        message = {
            "date": DATE_TEXT,
            "message_time": "2026-07-24 10:00:00",
            "send_time": 1,
            "conversation_name": "测试群",
            "sender": "李四",
            "type": "文本",
            "raw_text": "发生故障",
            "image_count_visible": 0,
            "files": [],
            "message_id": 101,
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            (day_dir / "raw_messages.jsonl").write_text(
                json.dumps(message, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
            write_issue_document(day_dir, issue_fields=fields, issues=[issue])
            outputs = export_day(day_dir, DATE_TEXT, "测试群")

            workbook = load_workbook(outputs["xlsx"], read_only=True, data_only=True)
            sheet = workbook["问题清单"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            workbook.close()
            markdown = Path(outputs["markdown"]).read_text(encoding="utf-8")

        self.assertEqual(
            headers,
            [
                "序号",
                "时间",
                "发送人",
                "严重级别",
                "标签",
                "影响数量",
                "已恢复",
                "截图引用",
                "问题Key",
            ],
        )
        self.assertEqual(values[3:7], ["P0", "客户、线上", "0", "否"])
        self.assertEqual(values[-1], "issue-custom-1")
        self.assertIn("- 严重级别：P0", markdown)
        self.assertIn("- 标签：客户、线上", markdown)
        self.assertIn("- 影响数量：0", markdown)
        self.assertIn("- 已恢复：否", markdown)
        self.assertNotIn("- 模块：", markdown)
        self.assertNotIn("- 问题分类：", markdown)


class SmartSheetTemplateTests(unittest.TestCase):
    def test_document_revision_is_canonical_and_binds_date_and_images(self):
        document = {
            "date": DATE_TEXT,
            "prompt": {"name": "提示词", "id": "prompt"},
            "issue_fields": [canonical_field("description", "问题描述", "text")],
            "image_manifest": {"bulk:101_01": "C:/work/capture.png"},
            "issues": [
                {"key": "issue-a", "values": {"description": "登录失败"}}
            ],
        }
        reordered = {
            "issues": copy.deepcopy(document["issues"]),
            "image_manifest": copy.deepcopy(document["image_manifest"]),
            "issue_fields": [dict(reversed(list(document["issue_fields"][0].items())))],
            "prompt": {"id": "prompt", "name": "提示词"},
            "date": DATE_TEXT,
        }
        revision = document_revision(document, DATE_TEXT)

        self.assertRegex(revision, r"^[0-9a-f]{64}$")
        self.assertEqual(document_revision(reordered, DATE_TEXT), revision)
        self.assertNotEqual(document_revision(document, "2026-07-25"), revision)
        changed_image = copy.deepcopy(document)
        changed_image["image_manifest"]["bulk:101_01"] = "C:/work/other.png"
        self.assertNotEqual(document_revision(changed_image, DATE_TEXT), revision)
        changed_issue = copy.deepcopy(document)
        changed_issue["issues"][0]["values"]["description"] = "支付失败"
        self.assertNotEqual(document_revision(changed_issue, DATE_TEXT), revision)

    def test_template_revision_is_canonical_and_changes_with_effective_config(self):
        template = smart_template("incident")
        reordered = {
            "field_mappings": copy.deepcopy(template["field_mappings"]),
            "schema": {
                key: dict(reversed(list(value.items())))
                for key, value in template["schema"].items()
            },
            "batch_size": template["batch_size"],
            "webhook_url_env": template["webhook_url_env"],
            "webhook_url": template["webhook_url"],
            "url": template["url"],
            "name": template["name"],
            "id": template["id"],
        }
        revision = template_revision(template)

        self.assertRegex(revision, r"^[0-9a-f]{64}$")
        self.assertEqual(template_revision(reordered), revision)
        self.assertNotIn(template["webhook_url"], revision)

        for key, replacement in {
            "id": "changed-id",
            "name": "changed-name",
            "url": "https://docs.qq.com/sheet/changed",
            "webhook_url": "https://example.invalid/changed-hook",
            "batch_size": 7,
        }.items():
            with self.subTest(key=key):
                changed = copy.deepcopy(template)
                changed[key] = replacement
                self.assertNotEqual(template_revision(changed), revision)

        changed_schema = copy.deepcopy(template)
        changed_schema["schema"]["f_desc"]["title"] = "*故障描述"
        self.assertNotEqual(template_revision(changed_schema), revision)
        changed_mapping = copy.deepcopy(template)
        changed_mapping["field_mappings"][0]["default_value"] = "默认描述"
        self.assertNotEqual(template_revision(changed_mapping), revision)

        env_template = smart_template(
            "env",
            webhook_url="",
            webhook_url_env="REVISION_TEST_HOOK",
        )
        with patch.dict(
            os.environ,
            {"REVISION_TEST_HOOK": "https://example.invalid/env-a"},
            clear=False,
        ):
            first_env_revision = template_revision(env_template)
        with patch.dict(
            os.environ,
            {"REVISION_TEST_HOOK": "https://example.invalid/env-b"},
            clear=False,
        ):
            self.assertNotEqual(template_revision(env_template), first_env_revision)

    def test_template_precedence_mapping_and_environment_webhook(self):
        config = {
            "smart_sheet": {
                "default_template_id": "global",
                "templates": [
                    smart_template("global", webhook_url="https://example.invalid/global"),
                    smart_template(
                        "bound",
                        webhook_url="",
                        webhook_url_env="CUSTOM_SMART_SHEET_HOOK",
                    ),
                    smart_template("explicit", webhook_url="https://example.invalid/explicit"),
                ],
            }
        }
        document = {"prompt": {"default_smart_sheet_template_id": "bound"}}
        bound = resolve_template(config, document)
        explicit = resolve_template(config, document, "explicit")
        self.assertEqual(bound["id"], "bound")
        self.assertEqual(explicit["id"], "explicit")
        with patch.dict(
            os.environ,
            {"CUSTOM_SMART_SHEET_HOOK": "https://example.invalid/from-env"},
            clear=False,
        ):
            self.assertEqual(
                resolve_webhook_url(bound),
                "https://example.invalid/from-env",
            )
        self.assertEqual(
            resolve_webhook_url(explicit),
            "https://example.invalid/explicit",
        )

        mapping_template = {
            "schema": {
                "f_severity": {
                    "title": "级别",
                    "type": "single_select",
                    "enum": ["P0", "P1"],
                },
                "f_count": {"title": "影响数", "type": "number"},
                "f_resolved": {"title": "已恢复", "type": "boolean"},
                "f_date": {"title": "登记日期", "type": "date_time"},
                "f_sender": {"title": "反馈人", "type": "text"},
            },
            "field_mappings": [
                {"source_key": "severity", "target_field_id": "f_severity", "target_type": "single_select"},
                {"source_key": "affected", "target_field_id": "f_count", "target_type": "number"},
                {"source_key": "resolved", "target_field_id": "f_resolved", "target_type": "boolean"},
                {"source_key": "$date", "target_field_id": "f_date", "target_type": "date_time"},
                {"source_key": "$sender", "target_field_id": "f_sender", "target_type": "text"},
            ],
        }
        issue = {
            "sender": "王五",
            "values": {"severity": "P0", "affected": 0, "resolved": False},
            "severity": "P1",
            "affected": 88,
            "resolved": True,
        }
        self.assertEqual(
            issue_values(mapping_template, issue, DATE_TEXT, []),
            {
                "f_severity": [{"text": "P0"}],
                "f_count": 0,
                "f_resolved": False,
                "f_date": date_time_millis(DATE_TEXT),
                "f_sender": "王五",
            },
        )

    def test_required_boolean_mapping_rejects_empty_but_preserves_explicit_false(self):
        fields = [canonical_field("resolved", "已恢复", "boolean")]
        template = smart_template("incident")
        template["schema"] = {"f_resolved": {"title": "*已恢复", "type": "boolean"}}
        template["field_mappings"] = [
            {
                "source_key": "resolved",
                "target_field_id": "f_resolved",
                "target_type": "boolean",
                "required": True,
                "default_value": "",
            }
        ]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            }
        }
        self.assertEqual(
            issue_values(template, {"values": {"resolved": ""}}, DATE_TEXT, []),
            {"f_resolved": ""},
        )
        self.assertEqual(
            issue_values(template, {"values": {"resolved": False}}, DATE_TEXT, []),
            {"f_resolved": False},
        )

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-empty", "values": {"resolved": ""}}],
            )
            with patch("worker.pipeline.smart_sheet.post_json") as post:
                with self.assertRaisesRegex(ValueError, r"必填字段.*已恢复.*为空"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=test_definition_path(day_dir),
                    )
                post.assert_not_called()

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-false", "values": {"resolved": False}}],
            )
            with patch(
                "worker.pipeline.smart_sheet.post_json",
                return_value={
                    "errcode": 0,
                    "add_records": [{"record_id": "remote-false"}],
                },
            ) as post:
                result = sync_issues(
                    config,
                    day_dir,
                    DATE_TEXT,
                    upload_images=False,
                    definition_path=test_definition_path(day_dir),
                )
            self.assertEqual(result["synced"], 1)
            self.assertIs(
                post.call_args.args[1]["add_records"][0]["values"]["f_resolved"],
                False,
            )

    def test_legacy_final_issues_file_exports_but_cannot_unlock_sync(self):
        template = smart_template("incident")
        template["field_mappings"][0]["source_key"] = "problem_description"
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            grouped = day_dir / "grouped_issues"
            grouped.mkdir(parents=True)
            legacy_path = grouped / "final_issues_20260724.json"
            legacy_path.write_text(
                json.dumps(
                    [
                        {
                            "key": "legacy-issue",
                            "问题描述": "旧问题文件中的支付失败",
                            "模块": "订单/售后",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (day_dir / "raw_messages.jsonl").write_text("", encoding="utf-8")
            outputs = export_day(
                day_dir,
                DATE_TEXT,
                "旧版群",
                export_xlsx=False,
                export_markdown=True,
            )
            markdown = Path(outputs["markdown"]).read_text(encoding="utf-8")
            self.assertIn("旧问题文件中的支付失败", markdown)

            with patch("worker.pipeline.smart_sheet.post_json") as post:
                with self.assertRaisesRegex(ValueError, "缺少不可变问题清单快照"):
                    preview_sync(config, day_dir, DATE_TEXT)
                with self.assertRaisesRegex(ValueError, "不可变问题清单快照"):
                    preview_sync(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=legacy_path,
                    )
                missing_snapshot = (
                    grouped / "snapshots" / "issue_definitions_20260724_missing.json"
                )
                with self.assertRaisesRegex(FileNotFoundError, "快照不存在"):
                    preview_sync(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=missing_snapshot,
                    )
                with self.assertRaisesRegex(ValueError, "不可变问题清单快照"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=legacy_path,
                    )
                post.assert_not_called()

    def test_naive_datetime_uses_shanghai_timezone_and_explicit_zone_is_preserved(self):
        shanghai = ZoneInfo("Asia/Shanghai")
        expected = int(
            datetime(2026, 7, 24, 10, 0, 0, tzinfo=shanghai).timestamp() * 1000
        )
        self.assertEqual(date_time_millis("2026-07-24 10:00:00"), str(expected))
        self.assertEqual(
            date_time_millis("2026-07-24T02:00:00Z"),
            str(expected),
        )
        self.assertNotEqual(date_time_millis(DATE_TEXT), utc_date_millis(DATE_TEXT))

    def test_old_ledger_migrates_only_to_legacy_default_and_templates_are_isolated(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        issue = {"key": "issue-a", "values": {"description": "登录失败"}}
        config = {
            "smart_sheet": {
                "default_template_id": "alpha",
                "templates": [smart_template("alpha"), smart_template("beta")],
            }
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[issue],
                prompt_template_id="beta",
            )
            ledger_path = day_dir / "smartsheet_desktop_sync_state.json"
            ledger_path.write_text(
                json.dumps(
                    {
                        "date": DATE_TEXT,
                        "updated_at": "2026-07-24T11:00:00+08:00",
                        "synced": {"issue-a": {"record_id": "legacy-record"}},
                    }
                ),
                encoding="utf-8",
            )

            migrated = load_ledger(ledger_path)
            self.assertEqual(migrated["version"], 2)
            self.assertEqual(
                template_synced(migrated, "default")["issue-a"]["record_id"],
                "legacy-record",
            )
            self.assertEqual(template_synced(migrated, "alpha"), {})
            self.assertEqual(template_synced(migrated, "beta"), {})
            template_synced(migrated, "beta")["issue-b"] = {"record_id": "beta-record"}
            save_ledger(ledger_path, migrated)
            reloaded = load_ledger(ledger_path)
            self.assertIn("issue-a", template_synced(reloaded, "default"))
            self.assertNotIn("issue-a", template_synced(reloaded, "alpha"))
            self.assertNotIn("issue-a", template_synced(reloaded, "beta"))
            self.assertIn("issue-b", template_synced(reloaded, "beta"))

            # Replace with the original ledger to verify preview-time migration as well.
            ledger_path.write_text(
                json.dumps({"synced": {"issue-a": {"record_id": "legacy-record"}}}),
                encoding="utf-8",
            )
            bound_preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=test_definition_path(day_dir),
            )
            alpha_preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                template_id="alpha",
                definition_path=test_definition_path(day_dir),
            )

        self.assertEqual(bound_preview["template_id"], "beta")
        self.assertEqual(bound_preview["pending"], 1)
        self.assertEqual(bound_preview["already_synced"], 0)
        self.assertEqual(alpha_preview["template_id"], "alpha")
        self.assertEqual(alpha_preview["pending"], 1)
        self.assertEqual(alpha_preview["already_synced"], 0)

    def test_invalid_or_unreadable_ledger_fails_closed_without_rewriting_it(self):
        with tempfile.TemporaryDirectory() as directory:
            ledger_path = Path(directory) / "smartsheet_desktop_sync_state.json"
            for label, original in {
                "invalid JSON": b'{"synced":',
                "non-object root": b"[]",
                "non-object templates": b'{"templates": []}',
                "non-object template state": b'{"templates": {"alpha": []}}',
                "missing synced state": b'{"templates": {"alpha": {}}}',
                "non-object synced state": b'{"templates": {"alpha": {"synced": []}}}',
            }.items():
                with self.subTest(label=label):
                    ledger_path.write_bytes(original)
                    with self.assertRaisesRegex(ValueError, "同步账本"):
                        load_ledger(ledger_path)
                    self.assertEqual(ledger_path.read_bytes(), original)

            original = b'{"synced": {}}'
            ledger_path.write_bytes(original)
            with patch.object(Path, "read_text", side_effect=OSError("access denied")):
                with self.assertRaisesRegex(ValueError, "无法读取同步账本"):
                    load_ledger(ledger_path)
            self.assertEqual(ledger_path.read_bytes(), original)

    def test_invalid_synced_entries_block_preview_and_sync_without_network_or_rewrite(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            }
        }
        invalid_ledgers = {
            "v2 null entry": {
                "version": 2,
                "templates": {
                    "incident": {"synced": {"issue-a": None}},
                },
            },
            "legacy missing record id": {
                "synced": {"issue-a": {}},
            },
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "登录失败"}}],
            )
            ledger_path = day_dir / "smartsheet_desktop_sync_state.json"
            for label, ledger in invalid_ledgers.items():
                with self.subTest(label=label):
                    original = json.dumps(ledger, ensure_ascii=False).encode("utf-8")
                    ledger_path.write_bytes(original)
                    with (
                        patch("worker.pipeline.smart_sheet.get_access_token") as token,
                        patch("worker.pipeline.smart_sheet.upload_image") as upload,
                        patch("worker.pipeline.smart_sheet.post_json") as post,
                    ):
                        with self.assertRaisesRegex(ValueError, "record_id"):
                            preview_sync(
                                config,
                                day_dir,
                                DATE_TEXT,
                                definition_path=test_definition_path(day_dir),
                            )
                        with self.assertRaisesRegex(ValueError, "record_id"):
                            sync_issues(
                                config,
                                day_dir,
                                DATE_TEXT,
                                upload_images=False,
                                definition_path=test_definition_path(day_dir),
                            )
                        token.assert_not_called()
                        upload.assert_not_called()
                        post.assert_not_called()
                    self.assertEqual(ledger_path.read_bytes(), original)

    def test_preview_revision_rejects_changed_config_before_network_calls(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "登录失败"}}],
            )
            preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=test_definition_path(day_dir),
            )
            self.assertEqual(
                preview["template_revision"],
                template_revision(config["smart_sheet"]["templates"][0]),
            )
            changed = copy.deepcopy(config)
            changed["smart_sheet"]["templates"][0]["batch_size"] = 7
            with (
                patch("worker.pipeline.smart_sheet.post_json") as post,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
            ):
                with self.assertRaisesRegex(ValueError, "配置已变化，请刷新预览后再确认"):
                    sync_issues(
                        changed,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=test_definition_path(day_dir),
                        expected_template_revision=preview["template_revision"],
                    )
                post.assert_not_called()
                upload.assert_not_called()

    def test_changed_definition_snapshot_is_rejected_before_network_calls(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            snapshot = write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "第一份"}}],
            )
            preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=snapshot,
            )
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "被修改"}}],
            )

            with (
                patch("worker.pipeline.smart_sheet.get_access_token") as token,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
                patch("worker.pipeline.smart_sheet.post_json") as post,
            ):
                with self.assertRaisesRegex(ValueError, "问题清单已变化，请刷新预览后再确认"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=snapshot,
                        expected_document_revision=preview["document_revision"],
                    )
                token.assert_not_called()
                upload.assert_not_called()
                post.assert_not_called()

    def test_concurrent_same_template_syncs_are_serialized_and_lock_recovers(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "登录失败"}}],
            )
            context = multiprocessing.get_context("spawn")
            first_ready = context.Event()
            second_ready = context.Event()
            entered = context.Event()
            release = context.Event()
            second_posted = context.Event()
            results = context.Queue()
            first = context.Process(
                target=run_sync_process,
                args=(
                    "first",
                    "incident",
                    True,
                    config,
                    str(day_dir),
                    first_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            second = context.Process(
                target=run_sync_process,
                args=(
                    "second",
                    "incident",
                    False,
                    config,
                    str(day_dir),
                    second_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            first.start()
            try:
                self.assertTrue(entered.wait(10), "first sync never reached POST")
                second.start()
                self.assertTrue(second_ready.wait(10), "second sync process never started")
                posted_while_first_owned_lock = second_posted.wait(1.5)
            finally:
                release.set()
                first.join(10)
                if second.pid is not None:
                    second.join(10)
                for process in (first, second):
                    if process.is_alive():
                        process.terminate()
                        process.join(5)

            self.assertFalse(posted_while_first_owned_lock)
            self.assertEqual(first.exitcode, 0)
            self.assertEqual(second.exitcode, 0)
            outcomes = {
                outcome["role"]: outcome
                for outcome in (results.get(timeout=5), results.get(timeout=5))
            }
            self.assertIn("intentional first-sync failure", outcomes["first"]["error"])
            self.assertEqual(outcomes["second"]["result"]["synced"], 1)
            self.assertTrue(second_posted.is_set())

    def test_concurrent_successful_same_template_sync_posts_only_once(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "登录失败"}}],
            )
            context = multiprocessing.get_context("spawn")
            first_ready = context.Event()
            second_ready = context.Event()
            entered = context.Event()
            release = context.Event()
            second_posted = context.Event()
            results = context.Queue()
            first = context.Process(
                target=run_sync_process,
                args=(
                    "first",
                    "incident",
                    False,
                    config,
                    str(day_dir),
                    first_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            second = context.Process(
                target=run_sync_process,
                args=(
                    "second",
                    "incident",
                    False,
                    config,
                    str(day_dir),
                    second_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            first.start()
            try:
                self.assertTrue(entered.wait(10), "first sync never reached POST")
                second.start()
                self.assertTrue(second_ready.wait(10), "second sync process never started")
                posted_while_first_owned_lock = second_posted.wait(1.5)
            finally:
                release.set()
                first.join(10)
                if second.pid is not None:
                    second.join(10)
                for process in (first, second):
                    if process.is_alive():
                        process.terminate()
                        process.join(5)

            self.assertFalse(posted_while_first_owned_lock)
            self.assertEqual(first.exitcode, 0)
            self.assertEqual(second.exitcode, 0)
            outcomes = {
                outcome["role"]: outcome
                for outcome in (results.get(timeout=5), results.get(timeout=5))
            }
            self.assertEqual(outcomes["first"]["result"]["synced"], 1)
            self.assertEqual(outcomes["second"]["result"]["synced"], 0)
            self.assertEqual(outcomes["second"]["result"]["skipped"], 1)
            self.assertFalse(second_posted.is_set())

    def test_concurrent_different_templates_preserve_both_ledger_states(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "smart_sheet": {
                "default_template_id": "alpha",
                "templates": [smart_template("alpha"), smart_template("beta")],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"description": "登录失败"}}],
            )
            context = multiprocessing.get_context("spawn")
            first_ready = context.Event()
            second_ready = context.Event()
            entered = context.Event()
            release = context.Event()
            second_posted = context.Event()
            results = context.Queue()
            first = context.Process(
                target=run_sync_process,
                args=(
                    "first",
                    "alpha",
                    False,
                    config,
                    str(day_dir),
                    first_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            second = context.Process(
                target=run_sync_process,
                args=(
                    "second",
                    "beta",
                    False,
                    config,
                    str(day_dir),
                    second_ready,
                    entered,
                    release,
                    second_posted,
                    results,
                ),
            )
            first.start()
            try:
                self.assertTrue(entered.wait(10), "first sync never reached POST")
                second.start()
                self.assertTrue(second_ready.wait(10), "second sync process never started")
                posted_while_first_owned_lock = second_posted.wait(1.5)
            finally:
                release.set()
                first.join(10)
                if second.pid is not None:
                    second.join(10)
                for process in (first, second):
                    if process.is_alive():
                        process.terminate()
                        process.join(5)

            self.assertFalse(posted_while_first_owned_lock)
            self.assertEqual(first.exitcode, 0)
            self.assertEqual(second.exitcode, 0)
            outcomes = {
                outcome["role"]: outcome
                for outcome in (results.get(timeout=5), results.get(timeout=5))
            }
            self.assertEqual(outcomes["first"]["result"]["synced"], 1)
            self.assertEqual(outcomes["second"]["result"]["synced"], 1)
            ledger = load_ledger(day_dir / "smartsheet_desktop_sync_state.json")
            self.assertIn("issue-a", template_synced(ledger, "alpha"))
            self.assertIn("issue-a", template_synced(ledger, "beta"))

    def test_issue_dedupe_survives_ordinal_changes_for_the_same_seed(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        config = {
            "config_version": 2,
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [smart_template("incident")],
            },
        }
        self.assertEqual(
            issue_dedupe_identity("issue_001_a1b2c3d4e5f6"),
            issue_dedupe_identity("issue_009_a1b2c3d4e5f6"),
        )
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue_009_a1b2c3d4e5f6",
                        "values": {"description": "同一来源消息"},
                    }
                ],
            )
            save_ledger(
                day_dir / "smartsheet_desktop_sync_state.json",
                {
                    "version": 2,
                    "templates": {
                        "incident": {
                            "synced": {
                                "issue_001_a1b2c3d4e5f6": {
                                    "record_id": "remote-existing",
                                }
                            }
                        }
                    },
                },
            )
            preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=test_definition_path(day_dir),
            )
            self.assertEqual(preview["already_synced"], 1)
            self.assertEqual(preview["pending"], 0)

    def test_sync_rejects_invalid_required_mapping_before_post_or_upload(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        invalid_template = smart_template("invalid")
        invalid_template["field_mappings"][0]["source_key"] = "missing_source"
        config = {
            "smart_sheet": {
                "default_template_id": "invalid",
                "templates": [invalid_template],
            }
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-a",
                        "values": {"description": "登录失败"},
                        "image_refs": ["bulk:101_01"],
                    }
                ],
            )
            with (
                patch("worker.pipeline.smart_sheet.post_json") as post,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
            ):
                with self.assertRaisesRegex(ValueError, "字段映射来源不存在"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=test_definition_path(day_dir),
                    )
                post.assert_not_called()
                upload.assert_not_called()

    def test_preview_rejects_incompatible_source_and_target_types(self):
        fields = [canonical_field("affected", "影响数", "text")]
        template = smart_template("incident")
        template["schema"] = {"f_count": {"title": "影响数", "type": "number"}}
        template["field_mappings"] = [
            {
                "source_key": "affected",
                "target_field_id": "f_count",
                "target_type": "number",
                "required": False,
                "default_value": "",
            }
        ]
        config = {
            "config_version": 2,
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            },
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-a", "values": {"affected": "很多"}}],
            )
            preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                definition_path=test_definition_path(day_dir),
            )
            self.assertFalse(preview["mapping_valid"])
            self.assertIn("字段映射类型不兼容", preview["validation_error"])
            with patch("worker.pipeline.smart_sheet.post_json") as post:
                with self.assertRaisesRegex(ValueError, "字段映射类型不兼容"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=test_definition_path(day_dir),
                    )
                post.assert_not_called()

    def test_sync_prevalidates_required_record_values_before_image_network_calls(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        template["schema"]["f_img"] = {"title": "截图", "type": "image"}
        template["field_mappings"].append(
            {
                "source_key": "$images",
                "target_field_id": "f_img",
                "target_type": "image",
                "required": False,
                "default_value": [],
            }
        )
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
                "upload": {"corpid": "corp-id", "corpsecret": "corp-secret"},
            }
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-empty-description",
                        "values": {"description": ""},
                        "image_refs": ["bulk:101_01"],
                    }
                ],
            )
            image_path = day_dir / "capture.png"
            image_path.write_bytes(
                b"\x89PNG\r\n\x1a\n" + b"\x00\x00\x00\rIHDR" + struct.pack(">II", 3, 2)
            )
            set_test_image_manifest(
                day_dir,
                {"bulk:101_01": str(image_path)},
            )
            manifest_dir = day_dir / "raw_attachments" / "_bulk_hd_cache"
            manifest_dir.mkdir(parents=True)
            (manifest_dir / "hd_cache_manifest.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "filename": "101_01_capture.png",
                                "local_path": str(image_path),
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            with (
                patch("worker.pipeline.smart_sheet.get_access_token", return_value="access-token") as token,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
                patch("worker.pipeline.smart_sheet.post_json") as post,
            ):
                with self.assertRaisesRegex(ValueError, "必填字段.*问题描述.*为空"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=test_definition_path(day_dir),
                    )
                token.assert_not_called()
                upload.assert_not_called()
                post.assert_not_called()

    def test_required_image_ref_without_manifest_fails_before_any_network_call(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        template["schema"]["f_img"] = {"title": "*截图", "type": "image"}
        template["field_mappings"].append(
            {
                "source_key": "$images",
                "target_field_id": "f_img",
                "target_type": "image",
                "required": True,
                "default_value": [],
            }
        )
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
                "upload": {"corpid": "corp-id", "corpsecret": "corp-secret"},
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-missing-image",
                        "values": {"description": "登录失败"},
                        "image_refs": ["bulk:101_01"],
                    }
                ],
            )
            with (
                patch("worker.pipeline.smart_sheet.get_access_token") as token,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
                patch("worker.pipeline.smart_sheet.post_json") as post,
            ):
                with self.assertRaisesRegex(ValueError, r"issue-missing-image.*截图.*本地文件"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=test_definition_path(day_dir),
                    )
                token.assert_not_called()
                upload.assert_not_called()
                post.assert_not_called()

    def test_all_pending_images_are_preflighted_before_uploading_the_first_record(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        template["schema"]["f_img"] = {"title": "*截图", "type": "image"}
        template["field_mappings"].append(
            {
                "source_key": "$images",
                "target_field_id": "f_img",
                "target_type": "image",
                "required": True,
                "default_value": [],
            }
        )
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
                "upload": {"corpid": "corp-id", "corpsecret": "corp-secret"},
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-with-image",
                        "values": {"description": "支付失败"},
                        "image_refs": ["bulk:101_01"],
                    },
                    {
                        "key": "issue-missing-image",
                        "values": {"description": "登录失败"},
                        "image_refs": ["bulk:102_01"],
                    },
                ],
            )
            image_path = day_dir / "capture.png"
            image_path.write_bytes(
                b"\x89PNG\r\n\x1a\n"
                + b"\x00\x00\x00\rIHDR"
                + struct.pack(">II", 3, 2)
            )
            set_test_image_manifest(
                day_dir,
                {"bulk:101_01": str(image_path)},
            )
            manifest_dir = day_dir / "raw_attachments" / "_bulk_hd_cache"
            manifest_dir.mkdir(parents=True)
            (manifest_dir / "hd_cache_manifest.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "filename": "101_01_capture.png",
                                "local_path": str(image_path),
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            with (
                patch("worker.pipeline.smart_sheet.get_access_token") as token,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
                patch("worker.pipeline.smart_sheet.post_json") as post,
            ):
                with self.assertRaisesRegex(ValueError, r"issue-missing-image.*截图.*本地文件"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=test_definition_path(day_dir),
                    )
                token.assert_not_called()
                upload.assert_not_called()
                post.assert_not_called()

    def test_corrupt_later_image_fails_before_token_upload_or_post(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        template["schema"]["f_img"] = {"title": "*截图", "type": "image"}
        template["field_mappings"].append(
            {
                "source_key": "$images",
                "target_field_id": "f_img",
                "target_type": "image",
                "required": True,
                "default_value": [],
            }
        )
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
                "upload": {"corpid": "corp-id", "corpsecret": "corp-secret"},
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-valid-image",
                        "values": {"description": "支付失败"},
                        "image_refs": ["bulk:101_01"],
                    },
                    {
                        "key": "issue-corrupt-image",
                        "values": {"description": "登录失败"},
                        "image_refs": ["bulk:102_01"],
                    },
                ],
            )
            valid_image = day_dir / "valid.png"
            valid_image.write_bytes(
                b"\x89PNG\r\n\x1a\n"
                + b"\x00\x00\x00\rIHDR"
                + struct.pack(">II", 3, 2)
            )
            corrupt_image = day_dir / "corrupt.bin"
            corrupt_image.write_bytes(b"not-an-image")
            set_test_image_manifest(
                day_dir,
                {
                    "bulk:101_01": str(valid_image),
                    "bulk:102_01": str(corrupt_image),
                },
            )

            with (
                patch("worker.pipeline.smart_sheet.get_access_token") as token,
                patch("worker.pipeline.smart_sheet.upload_image") as upload,
                patch("worker.pipeline.smart_sheet.post_json") as post,
            ):
                with self.assertRaisesRegex(
                    ValueError,
                    r"issue-corrupt-image.*无法识别有效尺寸",
                ):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        definition_path=test_definition_path(day_dir),
                    )
                token.assert_not_called()
                upload.assert_not_called()
                post.assert_not_called()

    def test_schema_required_marker_cannot_be_disabled_by_mapping_flag(self):
        fields = [canonical_field("description", "问题描述", "text")]
        template = smart_template("incident")
        template["field_mappings"][0]["required"] = False
        config = {
            "config_version": 2,
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            },
        }
        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[{"key": "issue-empty", "values": {"description": ""}}],
            )
            with patch("worker.pipeline.smart_sheet.post_json") as post:
                with self.assertRaisesRegex(ValueError, "必填字段.*问题描述.*为空"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=test_definition_path(day_dir),
                    )
                post.assert_not_called()

    def test_partial_success_persists_confirmed_ids_before_retry(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        config = {
            "config_version": 2,
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
            },
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {"key": "issue-a", "values": {"description": "支付失败"}},
                    {"key": "issue-b", "values": {"description": "登录失败"}},
                ],
            )
            with patch(
                "worker.pipeline.smart_sheet.post_json",
                return_value={
                    "errcode": 0,
                    "add_records": [{"record_id": "remote-record-a"}],
                },
            ) as post:
                with self.assertRaisesRegex(RuntimeError, "仅确认写入 1/2"):
                    sync_issues(
                        config,
                        day_dir,
                        DATE_TEXT,
                        upload_images=False,
                        definition_path=test_definition_path(day_dir),
                    )
                post.assert_called_once()

            ledger = load_ledger(
                day_dir / "smartsheet_desktop_sync_state.json",
            )
            synced = template_synced(ledger, "incident")
            self.assertEqual(synced["issue-a"]["record_id"], "remote-record-a")
            self.assertNotIn("issue-b", synced)
            preview = preview_sync(
                config,
                day_dir,
                DATE_TEXT,
                template_id="incident",
                definition_path=test_definition_path(day_dir),
            )
            self.assertEqual(preview["already_synced"], 1)
            self.assertEqual(preview["pending"], 1)

    def test_sync_uses_frozen_snapshot_images_after_live_manifest_is_overwritten(self):
        fields = [canonical_field("description", "问题描述", "text", required=True)]
        template = smart_template("incident")
        template["schema"]["f_img"] = {"title": "截图", "type": "image"}
        template["field_mappings"].append(
            {
                "source_key": "$images",
                "target_field_id": "f_img",
                "target_type": "image",
                "required": False,
                "default_value": [],
            }
        )
        config = {
            "smart_sheet": {
                "default_template_id": "incident",
                "templates": [template],
                "upload": {
                    "corpid": "corp-id",
                    "corpsecret": "corp-secret",
                    "delay_ms_between_image_uploads": 0,
                },
            }
        }

        with tempfile.TemporaryDirectory() as directory:
            day_dir = Path(directory)
            write_issue_document(
                day_dir,
                issue_fields=fields,
                issues=[
                    {
                        "key": "issue-a",
                        "sender": "赵六",
                        "values": {"description": "支付失败"},
                        "image_refs": ["bulk:101_01"],
                    }
                ],
            )
            image_path = day_dir / "capture.png"
            image_path.write_bytes(
                b"\x89PNG\r\n\x1a\n" + b"\x00\x00\x00\rIHDR" + struct.pack(">II", 3, 2)
            )
            set_test_image_manifest(
                day_dir,
                {"bulk:101_01": str(image_path)},
            )
            manifest_dir = day_dir / "raw_attachments" / "_bulk_hd_cache"
            manifest_dir.mkdir(parents=True)
            (manifest_dir / "hd_cache_manifest.json").write_text(
                json.dumps({"records": []}),
                encoding="utf-8",
            )

            with (
                patch("worker.pipeline.smart_sheet.get_access_token", return_value="access-token") as token,
                patch(
                    "worker.pipeline.smart_sheet.upload_image",
                    return_value="https://example.invalid/uploaded.png",
                ) as upload,
                patch(
                    "worker.pipeline.smart_sheet.post_json",
                    return_value={
                        "errcode": 0,
                        "add_records": [{"record_id": "remote-record-1"}],
                    },
                ) as post,
                patch("worker.pipeline.smart_sheet.time.sleep"),
            ):
                result = sync_issues(
                    config,
                    day_dir,
                    DATE_TEXT,
                    definition_path=test_definition_path(day_dir),
                )

            token.assert_called_once()
            upload.assert_called_once_with(
                config["smart_sheet"]["upload"],
                "access-token",
                image_path.resolve(),
            )
            post.assert_called_once()
            webhook_url, payload = post.call_args.args
            self.assertEqual(webhook_url, "https://example.invalid/hook")
            record_values = payload["add_records"][0]["values"]
            self.assertEqual(record_values["f_desc"], "支付失败")
            self.assertEqual(len(record_values["f_img"]), 1)
            uploaded_image = record_values["f_img"][0]
            self.assertEqual(uploaded_image["title"], "capture.png")
            self.assertEqual(
                uploaded_image["image_url"],
                "https://example.invalid/uploaded.png",
            )
            self.assertEqual((uploaded_image["width"], uploaded_image["height"]), (3, 2))
            self.assertTrue(uploaded_image["id"].startswith("desktop_001_01_"))
            self.assertEqual(result["synced"], 1)
            self.assertEqual(result["template_id"], "incident")

            ledger = load_ledger(
                day_dir / "smartsheet_desktop_sync_state.json",
            )
            self.assertEqual(
                template_synced(ledger, "incident")["issue-a"]["record_id"],
                "remote-record-1",
            )


if __name__ == "__main__":
    unittest.main()
